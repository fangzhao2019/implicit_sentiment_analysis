# -*- coding:utf-8 -*-
import jieba
import jieba.posseg as pseg
jieba.load_userdict('G:/userdict.txt')
import re
from openpyxl import load_workbook
from openpyxl import Workbook

featureSet=[]
wb=load_workbook('产品配置.xlsx')
ws=wb.active
for i in range(1,ws.max_row+1):
    feature=ws.cell(row=i,column=1).value
    featureSet.append(feature)
print(len(featureSet))
wb=load_workbook(u'1.1cutResult.xlsx')#待处理的数据
ws=wb.active
new_wb=Workbook()#创建新excel
new_ws=new_wb.active
new_ws.cell(row=1,column=1).value='comment'
new_ws.cell(row=1,column=2).value='score'
new_ws.cell(row=1,column=3).value='lenth'
new_ws.cell(row=1,column=4).value='featureNum'
new_ws.cell(row=1,column=5).value='feature'

m=1
for i in range(2,ws.max_row+1):
    if i%10000==0:print(i,m-1)
    comment=ws.cell(row=i,column=1).value
    score=ws.cell(row=i,column=2).value
    commentSet=[w for w in jieba.cut(comment)]
    #特征统计
    featureCount=[f for f in featureSet if f in commentSet]
    if len(featureCount)>0:#提取含有特征的句子
        count=len(featureCount)
        feature=','.join(featureCount)
        m+=1
        new_ws.cell(row=m,column=1).value=comment
        new_ws.cell(row=m,column=2).value=score
        new_ws.cell(row=m,column=3).value=len(commentSet)
        new_ws.cell(row=m,column=4).value=count
        new_ws.cell(row=m,column=5).value=feature
new_wb.save(u'commentSet.xlsx')

