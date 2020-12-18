# -*- coding: utf-8 -*-
from __future__ import division
import jieba.posseg as pseg
from openpyxl import load_workbook
from openpyxl import Workbook
import apriori
import time


def load_dataSet(filename):#加载excel，获取数据
    wb=load_workbook(filename)
    ws=wb.active
    dataSet=[]
    for i in range(2,ws.max_row+1):
        data=ws.cell(row=i,column=1).value
        dataSet.append(data)
    print(u'正在获取%s的评论数据'%filename)
    print("*"*80)
    return dataSet

def process_data(dataSet):#获取单个特征评论数据集，经过处理后返回满足最小支持度得频繁项（一项和二项）以及其支持度
    D=[]
    for i in range(len(dataSet)):
        data=[]
        words=pseg.cut(dataSet[i])
        for w in words:
            if w.flag=='n':
                data.append(w.word)
        D.append(data)
    L,suppData=apriori.apriori(D,100/len(D))
    resultSet=[]
    for i in range(len(L)):
        for j in range(len(L[i])):
            ss=list(L[i][j])
            ss.append(suppData[L[i][j]])
            resultSet.append(ss)    
    return resultSet

def process_all_data(filename1,filename2):#处理所有特征评论数据，并将结果存储到一张新的excel
    new_wb=Workbook()#创建存储结果的excel
    new_ws=new_wb.active
    dataSet=load_dataSet(filename1)
    m=len(dataSet)
    print(u'数据一共有%d个'%m)
    resultSet=process_data(dataSet)
    new_ws.cell(row=1,column=1).value=u'特征'
    new_ws.cell(row=1,column=2).value=u'支持度'
    for k in range(2,len(resultSet)+2):
        new_ws.cell(row=k,column=1).value=",".join(resultSet[k-2][:-1])
        new_ws.cell(row=k,column=2).value=resultSet[k-2][-1]
    new_wb.save(filename2)
    print('\n')
process_all_data(u'1.1cutResult.xlsx',u'resul4.xlsx')


