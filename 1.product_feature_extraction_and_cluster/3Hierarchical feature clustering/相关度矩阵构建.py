# -*- coding:utf-8 -*-
from __future__ import division
#######载入外部词典
import jieba
jieba.load_userdict('G:/userdict.txt')
import copy
from openpyxl import load_workbook
from openpyxl import Workbook
import numpy
from numpy import log

def load_featureSet(filename):#载入数据集
    wb=load_workbook(filename)
    ws=wb.active
    featureSet=[]
    for i in range(1,ws.max_row+1):
        feature=ws.cell(row=i,column=1).value        
        featureSet.append(feature)
    return featureSet

def calPMImat(filename,featureSet):#特征相关度计算
    size=len(featureSet)#特征数量
    featureFreq=numpy.ones(size)#存储特征频率
    PMImat=numpy.ones((size,size))#存储特征之间的关系

    wb=load_workbook(filename)
    ws=wb.active#载入数据文件
    commentNumber=ws.max_row-1

    for i in range(2,ws.max_row+1):
        if i%10000==0:print(i)#每隔10000行数据计数一次
        comment=ws.cell(row=i,column=1).value
        commentVec=[w for w in jieba.cut(comment)]#词向量

        featureExist=[]#存储当前评论出现的特征数量
        for j in range(len(featureSet)):
            if featureSet[j] in commentVec:
                featureFreq[j]+=1#特征数量统计
                featureExist.append(featureSet[j])
        if len(featureExist)==0:continue#若无特征，则不计算
        for feature1 in featureExist:#统计每两个特征共同出现的次数
            for feature2 in featureExist:
                m=featureSet.index(feature1)
                n=featureSet.index(feature2)
                PMImat[m,n]+=1
    for i in range(size):#根据上述结果计算PMI距离矩阵
        for j in range(size):
            Pi=featureFreq[i]/commentNumber
            Pj=featureFreq[j]/commentNumber
            Pij=PMImat[i][j]/commentNumber
            PMImat[i][j]=log(Pij/(Pi*Pj))/log(2)
    pmax=PMImat.max()
    pmin=PMImat.min()
    for i in range(size):
        for j in range(size):
            p=PMImat[i][j]
            PMImat[i][j]=(pmax-p)/(pmax-pmin)#############对矩阵进行标准化
    return featureFreq,PMImat

def createDataSet(featureSet):#初始化数据格式——二元列表
    dataSet=[[feature] for feature in featureSet]
    return dataSet

def distLinkage(cluster1,cluster2,featureSet,PMImat,linkage):
#计算两个簇之间的距离,linkage表示距离度量方式
    distance=[]
    for f1 in cluster1:
        for f2 in cluster2:
            m=featureSet.index(f1)
            n=featureSet.index(f2)
            distance.append(PMImat[m,n])
    if linkage=='singleLinkage':return max(distance)
    if linkage=='completeLinkage':return min(distance)
    if linkage=='averageLinkage':return sum(distance)/len(distance)

def agglomerativeClustering(dataSet,featureSet,PMImat,n_cluster,linkage):
#层次聚类法，输入为二元列表，n_cluster表示聚类个数
    clusterSet=copy.deepcopy(dataSet)
    n=len(clusterSet)#类别数量
    while n>n_cluster:
        #计算每两个类簇之间的距离,一个储存距离，一个储存相比较的簇的索引
        distanceIndex=[]
        distanceList=[]
        for i in range(len(clusterSet)):
            for j in range(i+1,len(clusterSet)):
                distance=distLinkage(clusterSet[i],clusterSet[j],featureSet,PMImat,linkage)
                distanceIndex.append([i,j])
                distanceList.append(distance)
        #找到距离最小的两个类簇(最小距离=最大相关度)
        minDistance=max(distanceList)#获取最大相关度
        minIndex=distanceList.index(minDistance)#获取最大相关度的位置
        x,y=distanceIndex[minIndex]#获取产生最大相关度的两个比较对象的索引
        #合并最近的两个簇，得到新的类簇数据集
        clusterSet[x].extend(clusterSet[y])
        del clusterSet[y]
        n=n-1 
    return clusterSet

def savePMImat1(featureSet,PMImat):
    wb=Workbook()
    ws=wb.active
    for i in range(len(featureSet)):
        ws.cell(row=i+2,column=1).value=featureSet[i]
        ws.cell(row=1,column=i+2).value=featureSet[i]
        for j in range(len(featureSet)):
            ws.cell(row=i+2,column=j+2).value=PMImat[i][j]
    wb.save('PMImat.xlsx')

def savePMImat2(filename,featureSet,PMImat):
    f=open(filename,'w+')
    f.write(','.join(featureSet))
    for i in range(len(featureSet)):
        f.write('\n')
        mmm=[str(a) for a in PMImat[i]]
        f.write(','.join(mmm))
    f.close()

featureSet=load_featureSet('../featureSet.xlsx')#特征集
#featureSet=[u'内饰',u'做工',u'用料',u'仪表盘',u'中控',u'材料']
featureFreq,PMImat=calPMImat('commentSet2.xlsx',featureSet)
savePMImat2('PMImat.txt',featureSet,PMImat)

#dataSet=createDataSet(featureSet)
#clusterResult1=agglomerativeClustering(dataSet,featureSet,PMImat,8,'singleLinkage')
#clusterResult2=agglomerativeClustering(dataSet,featureSet,PMImat,16,'completeLinkage')
#clusterResult3=agglomerativeClustering(dataSet,featureSet,PMImat,16,'averageLinkage')
