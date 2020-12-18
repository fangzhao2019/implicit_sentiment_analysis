# coding:utf-8
from __future__ import division
import jieba
jieba.load_userdict('G:/userdict.txt')
from openpyxl import load_workbook
from openpyxl import Workbook
import operator
from numpy import log
import time
time1=time.time()

def load_synFeatureSet():
    f=open('同义关系.txt',encoding='utf-8')
    dataSet=[]
    for word in f.readlines():
        dataSet.append(word.strip().split('、'))
    return dataSet

def convertSyn(word,synFeatureSet):
    for syn in synFeatureSet:
        if word in syn:
            return syn[0]
    return word

synFeatureSet=load_synFeatureSet()

wb=load_workbook('commentSet.xlsx')
ws=wb.active
for i in range(2,ws.max_row+1):
    comment=ws.cell(row=i,column=1).value
    commentVec=[w for w in jieba.cut(comment)]
    for j in range(len(commentVec)):
        commentVec[j]=convertSyn(commentVec[j],synFeatureSet)
    ws.cell(row=i,column=1).value=''.join(commentVec)

wb.save('commentSet2.xlsx')

time2=time.time()
print('一共耗时%d秒'%(time2-time1))
