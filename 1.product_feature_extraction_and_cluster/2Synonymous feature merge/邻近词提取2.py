# coding:utf-8
from __future__ import division
import jieba
jieba.load_userdict('G:/userdict.txt')
from openpyxl import load_workbook
from openpyxl import Workbook
import operator
from numpy import log
import time
time1=time.time()

def load_featureSet():
    featureSet=[]
    wb=load_workbook('../1初始特征提取/产品配置.xlsx')
    ws=wb.active
    for i in range(1,ws.max_row+1):
        feature=ws.cell(row=i,column=1).value
        featureSet.append(feature)
    return featureSet
featureSet=load_featureSet()

def load_commentVecSet():
    wb=load_workbook(u'commentSet.xlsx')##载入数据
    ws=wb.active
    commentVecSet=[]
    wordFrequencyDic={}#初始化词频词典
    for feature in featureSet:
        if not feature in wordFrequencyDic.keys():
            wordFrequencyDic[feature]=1
            
    for i in range(2,ws.max_row+1):
        comment=ws.cell(row=i,column=1).value
        commentVec=[w for w in jieba.cut(comment)]
        for word in list(set(commentVec)):
            if not word in wordFrequencyDic.keys():
                wordFrequencyDic[word]=0
            wordFrequencyDic[word]+=1
        commentVecSet.append(commentVec)

    for i in range(len(commentVecSet)):
        commentVecSet[i]=[w for w in commentVecSet[i] if wordFrequencyDic[w]>10]##########过滤低频词
    return commentVecSet,wordFrequencyDic
commentVecSet,wordFrequencyDic=load_commentVecSet()

def myfind(x,y):
    return [ a for a in range(len(y)) if y[a] == x]

def abs(x):
    if x<0:return -x
    else:return x

#找出特征词的左右显著邻近词
def calSNWmat(featureSet,commentVecSet,wordFrequencyDic):
    M=len(commentVecSet)
    
    distanceMat=[]#初始化距离矩阵
    for i in range(len(featureSet)):
        a={}
        for l in featureSet:
            if not l in a.keys():
                a[l]=1
        b={}
        for r in featureSet:
            if not r in b.keys():
                b[r]=1
        distanceMat.append([a,b])
        
    SNWmat=[]#初始化邻近词矩阵
    for i in range(len(featureSet)):
        SNWmat.append([{},{}])

    #首先，计算特征词的左右词词频
    for i in range(len(commentVecSet)):
        comment=commentVecSet[i]

        for j in range(len(featureSet)):
            feature=featureSet[j]
            f_index=myfind(feature,comment)
            for index in f_index:
                for k in range(1,len(comment)):###############距离5以内的词
                    if index-k>=0:
                        w_before=comment[index-k]
                        if not w_before in distanceMat[j][0].keys():
                            distanceMat[j][0][w_before]=1
                        distanceMat[j][0][w_before]+=1
                    if index+k<len(comment):
                        w_after=comment[index+k]
                        if not w_after in distanceMat[j][1].keys():
                            distanceMat[j][1][w_after]=1
                        distanceMat[j][1][w_after]+=1
    #然后，计算左右词PMI值
    for i in range(len(distanceMat)):
        for j in range(len(distanceMat[i])):
            for key in distanceMat[i][j].keys():
                count_wv=distanceMat[i][j][key]
                count_w=wordFrequencyDic[key]
                count_v=wordFrequencyDic[featureSet[i]]
                distanceMat[i][j][key]=abs(log(count_wv*M/(count_w*count_v))/log(2))
    #最后，排序并选择z个显著邻近词

    for i in range(len(distanceMat)):
        for j in range(len(distanceMat[i])):
            b=sorted(distanceMat[i][j].items(),key=operator.itemgetter(1),reverse=True)
            SNWmat[i][j]=[ww[0] for ww in b[:50]]##############z值的设定
            
    return SNWmat,distanceMat


#计算上下文相关性
def calFeatureSimilarity(featureSet,SNWmat,distanceMat):
    FSmat=[]
    for i in range(len(featureSet)):
        for j in range(len(featureSet)):
            f1=featureSet[i]
            f2=featureSet[j]
            f1_left=SNWmat[i][0]
            f1_right=SNWmat[i][1]
            f2_left=SNWmat[j][0]
            f2_right=SNWmat[j][1]

            f_left=list(set(f1_left)&set(f2_left))
            f_right=list(set(f1_right)&set(f2_right))

            FS=0
            for word in f_left:
                FS+=distanceMat[i][0][word]
                FS+=distanceMat[j][0][word]
            for word in f_right:
                FS+=distanceMat[i][1][word]
                FS+=distanceMat[j][1][word]
            den=1#添加这两个特征共现的分母
            den+=distanceMat[i][0][f2]
            den+=distanceMat[i][1][f2]
            den+=distanceMat[j][0][f1]
            den+=distanceMat[j][1][f1]

            FS=FS/(10*den)
            
            FSmat.append([f1,f2,FS])
    wb=Workbook()
    ws=wb.active
    for i in range(len(FSmat)):
        ws.cell(row=i+1,column=1).value=FSmat[i][0]
        ws.cell(row=i+1,column=2).value=FSmat[i][1]
        ws.cell(row=i+1,column=3).value=FSmat[i][2]
    wb.save('FS.xlsx')
    
SNWmat,distanceMat=calSNWmat(featureSet,commentVecSet,wordFrequencyDic)
calFeatureSimilarity(featureSet,SNWmat,distanceMat)
#b=sorted(Dic.items(),key=operator.itemgetter(1),reverse=True)

time2=time.time()
print('一共耗时%d秒'%(time2-time1))
