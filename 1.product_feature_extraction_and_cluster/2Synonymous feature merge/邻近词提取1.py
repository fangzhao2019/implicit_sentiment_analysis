# coding:utf-8
from __future__ import division
import jieba
jieba.load_userdict('G:/userdict.txt')
from openpyxl import load_workbook
from openpyxl import Workbook
import operator
from numpy import log

def load_featureSet():
    featureSet=[]
    wb=load_workbook('../产品配置.xlsx')
    ws=wb.active
    for i in range(1,ws.max_row+1):
        feature=ws.cell(row=i,column=1).value
        featureSet.append(feature)
    return featureSet
featureSet=load_featureSet()

def load_commentVecSet():
    #filterWord=['，','。']
    wb=load_workbook(u'commentSet.xlsx')################
    ws=wb.active
    commentVecSet=[]
    for i in range(2,ws.max_row+1):
        comment=ws.cell(row=i,column=1).value
        commentVec=[w for w in jieba.cut(comment)]
        commentVecSet.append(commentVec)
    return commentVecSet
commentVecSet=load_commentVecSet()

def myfind(x,y):
    return [ a for a in range(len(y)) if y[a] == x]

#找出特征词的左右显著邻近词
def calSNWmat(featureSet,commentVecSet):
    M=len(commentVecSet)
    wordFrequencyDic={}#初始化词频词典
    for feature in featureSet:
        if not feature in wordFrequencyDic.keys():
            wordFrequencyDic[feature]=1
    SNWmat=[]#初始化邻近词矩阵
    for i in range(len(featureSet)):
        SNWmat.append([{},{}])

    #首先，计算特征词的左右词词频
    for i in range(len(commentVecSet)):#对于每条评论
        comment=commentVecSet[i]
        
        for c in list(set(comment)):#统计词频
            if not c in wordFrequencyDic.keys():
                wordFrequencyDic[c]=1
            wordFrequencyDic[c]+=1
            
        for j in range(len(featureSet)):#对于每个特征
            feature=featureSet[j]
            f_index=myfind(feature,comment)#获取特征的索引
            if len(f_index)>0:index=f_index[0]
            else:continue
            for k in range(1,len(comment)):#距离5以内的词
                if index-k>=0:
                    w_before=comment[index-k]
                    if not w_before in SNWmat[j][0].keys():
                        SNWmat[j][0][w_before]=1
                    SNWmat[j][0][w_before]+=1
                if index+k<len(comment):
                    w_after=comment[index+k]
                    if not w_after in SNWmat[j][1].keys():
                        SNWmat[j][1][w_after]=1
                    SNWmat[j][1][w_after]+=1
    #然后，计算左右词PMI值
    for i in range(len(SNWmat)):
        for j in range(len(SNWmat[i])):
            for key in SNWmat[i][j].keys():
                count_wv=SNWmat[i][j][key]
                count_w=wordFrequencyDic[key]
                count_v=wordFrequencyDic[featureSet[i]]
                SNWmat[i][j][key]=log(count_wv*M/(count_w*count_v))/log(2)
    #最后，排序并选择z个显著邻近词
    left_word=[]
    right_word=[]
    for i in range(len(SNWmat)):
        for j in range(len(SNWmat[i])):
            b=sorted(SNWmat[i][j].items(),key=operator.itemgetter(1),reverse=True)
            SNWmat[i][j]=b[:20]#z值的设定
            wordList=[word[0] for word in b[:20]]
            if j==0:
                left_word.extend(wordList)
            if j==1:
                right_word.extend(wordList)
    return SNWmat,list(set(left_word)),list(set(right_word)),wordFrequencyDic

#计算每个特征词与所有邻近词的距离
def createDistanceMat(featureSet,commentVecSet,left_word,right_word,wordFrequencyDic):
    M=len(commentVecSet)
    distanceMat=[]
    left_word=left_word+featureSet
    right_word=right_word+featureSet
    for i in range(len(featureSet)):
        a={}
        for l in left_word:
            if not l in a.keys():
                a[l]=1
        b={}
        for r in right_word:
            if not r in b.keys():
                b[r]=1
        distanceMat.append([a,b])
        
    for i in range(len(commentVecSet)):
        comment=commentVecSet[i]
        for j in range(len(featureSet)):
            feature=featureSet[j]
            f_index=myfind(feature,comment)#获取特征的索引
            if len(f_index)>0:index=f_index[0]
            else:continue
            for k in range(1,len(comment)):
                if index-k>=0:
                    w_before=comment[index-k]
                    if w_before in left_word:
                        distanceMat[j][0][w_before]+=1
                if index+k<len(comment):
                    w_after=comment[index+k]
                    if w_after in right_word:
                        distanceMat[j][1][w_after]+=1
                        
    for i in range(len(distanceMat)):
        for j in range(len(distanceMat[i])):
            for key in distanceMat[i][j].keys():
                count_wv=distanceMat[i][j][key]
                count_w=wordFrequencyDic[key]
                count_v=wordFrequencyDic[featureSet[i]]
                distanceMat[i][j][key]=log(count_wv*M/(count_w*count_v))/log(2)
    return distanceMat

#计算上下文相关性
def calFeatureSimilarity(featureSet,SNWmat,distanceMat):
    FSmat=[]
    for i in range(len(featureSet)):
        for j in range(i+1,len(featureSet)):
            f1=featureSet[i]
            f2=featureSet[j]
            f1_left=[w[0] for w in SNWmat[i][0]]
            f1_right=[w[0] for w in SNWmat[i][1]]
            f2_left=[w[0] for w in SNWmat[j][0]]
            f2_right=[w[0] for w in SNWmat[j][1]]

            f_left=f1_left+f2_left
            f_right=f1_right+f2_right

            FS=0
            for word in f_left:
                FS+=distanceMat[i][0][word]
                FS+=distanceMat[j][0][word]
            for word in f_right:
                FS+=distanceMat[i][1][word]
                FS+=distanceMat[j][1][word]
            den=1#添加这两个特征共现的分母
            den+=distanceMat[i][0][f2]
            den+=distanceMat[i][1][f2]
            den+=distanceMat[j][0][f1]
            den+=distanceMat[j][1][f1]

            FS=FS/(10*den)
            
            FSmat.append([f1,f2,FS])
    wb=Workbook()
    ws=wb.active
    for i in range(len(FSmat)):
        ws.cell(row=i+1,column=1).value=FSmat[i][0]
        ws.cell(row=i+1,column=2).value=FSmat[i][1]
        ws.cell(row=i+1,column=3).value=FSmat[i][2]
    wb.save('FS.xlsx')
    
SNWmat,left_word,right_word,wordFrequencyDic=calSNWmat(featureSet,commentVecSet)
distanceMat=createDistanceMat(featureSet,commentVecSet,left_word,right_word,wordFrequencyDic)
calFeatureSimilarity(featureSet,SNWmat,distanceMat)
#b=sorted(Dic.items(),key=operator.itemgetter(1),reverse=True)
