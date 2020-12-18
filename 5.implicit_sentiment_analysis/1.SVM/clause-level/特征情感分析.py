# coding:utf-8
from __future__ import division
from openpyxl import load_workbook
import numpy as np
import jieba
import jieba.posseg as pseg
jieba.load_userdict('G:/userdict.txt')
import os
import time
from sklearn import svm
from sklearn.naive_bayes import BernoulliNB

#载入txt数据集，如特征,停用词
def load_dataSet(filename):
    f=open(filename,encoding='utf-8')
    dataSet=[]
    for word in f.readlines():
        dataSet.append(word.strip())
    return dataSet

#导入标记好的初始评论数据
def load_initComment(filename):
    wb=load_workbook(filename)
    ws=wb.active
    commentSet=[]
    labels=[]
    for i in range(2,ws.max_row+1):
        comment=ws.cell(row=i,column=1).value
        label=int(ws.cell(row=i,column=2).value)
        commentSet.append(comment)
        labels.append(label)
    return np.array(commentSet),np.array(labels)

#对评论进行分词，筛选掉停用词，存储为（word,pos）对的集合
def commentPOS(commentSet,stopWords,method='jieba'):
    commentVecSet=[]
    if method=='jieba':
        for i in range(len(commentSet)):
            commentVec=[(w.word,w.flag) for w in pseg.cut(commentSet[i]) if w.word not in stopWords]
            commentVecSet.append(commentVec)
    return np.array(commentVecSet)

#将分词后的评论向量根据词性进行筛选
def commentVecSet_Filter(commentVecSet,POS):
    commentPOSVec=[]
    for comment in commentVecSet:
        if POS=='all':
            POSVec=[com[0] for com in comment]
        else:
            POSVec=[com[0] for com in comment if com[1] in POS]
        commentPOSVec.append(POSVec)
    return commentPOSVec

#若参数为vocab，统计每个元素出现的频率，存储为一个词典
#若参数为gram，统计每个2-POS-gram出现的频率，存储为一个词典
def createVocabList(commentPOSVec,types='vocab'):
    vocabList={}
    for comment in commentPOSVec:
        if types=='gram':
            commentVec=[(comment[i],comment[i+1]) for i in range(len(comment)-1)]
        if types=='vocab':
            commentVec=comment
        for word in commentVec:
            if not word in vocabList.keys():
                vocabList[word]=0
            vocabList[word]+=1
    return vocabList

#若参数为vocab，根据出现频率对词汇进行筛选
#若参数为gram，根据出现频率对词对进行筛选
def vocabFilter(vocabList,freq,types='vocab'):
    myVocab=[]
    sortResult=sorted(vocabList.items(),key=lambda e:e[1],reverse=True)
    for word in sortResult:
        if word[1]>freq:
            if types=='vocab':
                myVocab.append(word[0])
            if types=='gram':
                myVocab.append(word[0])
    return myVocab

def load_patternSet(filename):
    wb=load_workbook(filename)
    ws=wb.active
    
    patternSet=[]
    for i in range(2,ws.max_row+1):
        pattern=ws.cell(row=i,column=1).value.split(',')
        patternSet.append(pattern)
    return patternSet

#判断一个模式是否为另一个模式的子模式
def pattern_contain(pattern,comment):
    all_index=[]
    #首先计算出待测模式中的每个词在评论中的位置
    #print(pattern)
    for pat in pattern:
        all_index.append([i for i in range(len(comment)) if comment[i]==pat])
    position=-1
    for i in range(len(all_index)):
        if len(all_index[i])==0:return 0
        count=0
        for j in range(len(all_index[i])):
            if all_index[i][j]>position:
                position=all_index[i][j]
                break
            else:
                count+=1
        if count==len(all_index[i]):
            return 0
    return 1

#为输入文本构建分类属性向量，包含为1，否则为0
def setOfWords2Vec(vocabList, inputSet):
    returnVec = [0]*len(vocabList)
    for word in inputSet:
        if word in vocabList:
            returnVec[vocabList.index(word)] = 1
    return returnVec

def setOfGrams2Vec(gramList, inputSet):
    returnVec=[0]*len(gramList)
    for pattern in gramList:
        if pattern_contain(pattern,inputSet):
            returnVec[gramList.index(pattern)]=1
    return returnVec

def create_dataMat2(commentPOSVec,myVocab,myPattern):
    dataMat=[]
    for i in range(len(commentPOSVec)):
        data=setOfWords2Vec(myVocab, commentPOSVec[i])
        #data.extend(setOfGrams2Vec(myPattern,commentPOSVec[i]))
        dataMat.append(data)
    return np.array(dataMat)

time1=time.time()
##########################################################################################
#训练数据
stopWords=u'，。的 ；了、：吧'
commentSet,trainLabel=load_initComment('clauseSet.xlsx')
commentVecSet=commentPOS(commentSet,stopWords,method='jieba')
commentPOSVec=commentVecSet_Filter(commentVecSet,['n','a','d','v'])

vocabList=createVocabList(commentPOSVec,'vocab')
myVocab=vocabFilter(vocabList,5,'vocab')
print('共有单词%d个'%(len(myVocab)))
    
myPattern=load_patternSet(u'frequentPatternSet.xlsx')
print('共有模式%d个'%(len(myPattern)))
#选择特征构建数据矩阵
trainMat=create_dataMat2(commentPOSVec,myVocab,myPattern)
print(u'数据矩阵构建完毕')

#训练模型

clf=BernoulliNB()
#clf=svm.SVC(C=1,kernel='linear')
clf=clf.fit(trainMat,trainLabel)

wb=load_workbook('clauseTest.xlsx')
ws=wb.active
for i in range(2,ws.max_row+1):
    comment=ws.cell(row=i,column=1).value
    commentVec=[w.word for w in pseg.cut(comment)]
    testData=setOfWords2Vec(myVocab,commentVec)
    #testData.extend(setOfGrams2Vec(myPattern,commentVec))
    predictor=clf.predict([testData])[0]
    ws.cell(row=i,column=8).value=predictor
wb.save('clauseResult.xlsx')
time2=time.time()
print('耗时%d秒'%(time2-time1))
