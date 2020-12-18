# coding:utf-8
from __future__ import division
from openpyxl import load_workbook
import numpy as np
import jieba
import jieba.posseg as pseg
jieba.load_userdict('G:/userdict.txt')
import os
import time
import gensim
from keras.preprocessing import sequence
from keras.models import Sequential
from keras.layers import Convolution2D, MaxPooling2D
from keras.layers.core import Dense, Dropout, Activation, Flatten

#导入标记好的初始评论数据
def load_initComment(filename):
    wb=load_workbook(filename)
    ws=wb.active
    commentSet=[]
    labels=[]
    for i in range(2,ws.max_row+1):
        comment=ws.cell(row=i,column=1).value
        label=int(ws.cell(row=i,column=2).value)
        commentVec=[w.word for w in pseg.cut(comment)]
        commentSet.append(comment)
        labels.append(label)
    return np.array(commentSet),np.array(labels)

def load_testComment(filename):
    wb=load_workbook(filename)
    ws=wb.active
    commentSet=[]
    labels=[]
    for i in range(2,ws.max_row+1):
        comment=ws.cell(row=i,column=1).value
        label=int(ws.cell(row=i,column=2).value)
        commentVec=[w.word for w in pseg.cut(comment)]
        commentSet.append(comment)
        labels.append(label)
    return np.array(commentSet),np.array(labels)

#为输入文本构建词向量
def create_vector(sen,padding_size,vec_size):
    matrix=[]
    for i in range(padding_size):
        try:
            matrix.append(model[sen[i]].tolist())
        except:
                # 这里有两种except情况，
                # 1. 这个单词找不到
                # 2. sen没那么长
                # 不管哪种情况，我们直接贴上全是0的vec
            matrix.append([0] * vec_size)
    return matrix

def transform_to_matrix(x,model, padding_size=256, vec_size=128):
    res = []
    for sen in x:
        matrix =create_vector(sen,padding_size,vec_size)  
        res.append(matrix)
    return res

time1=time.time()

X_commentSet,y_train=load_initComment('reviewSet.xlsx')
Y_commentSet,y_test=load_initComment('reviewTest.xlsx')
model=gensim.models.Word2Vec.load('word2vec/carCommentData.model')
#选择特征构建数据矩阵
trainMat=transform_to_matrix(X_commentSet,model)
testMat=transform_to_matrix(Y_commentSet,model)
print(u'数据矩阵构建完毕')
np.save('Y_train.npy',y_train)
np.save('X_train.npy',trainMat)

np.save('Y_test.npy',y_test)
np.save('X_test.npy',testMat)

time2=time.time()
print('耗时%d秒'%(time2-time1))
