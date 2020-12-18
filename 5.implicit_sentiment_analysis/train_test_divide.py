# -*- coding:utf-8 -*-
from __future__ import division
import jieba
import jieba.posseg as pseg
jieba.load_userdict('G:/userdict.txt')
from openpyxl import load_workbook
from openpyxl import Workbook

def load_featureCategory():
    f=open('featureCategory.txt',encoding='utf-8')
    dataSet=[]
    for w in f.readlines():
        dataSet.append(w.replace('\n','').split('、'))
    return dataSet[1:]

def featureReplace(f,featureCategory):
    for feature in featureCategory:
        if f in feature:
            return feature[0]

featureCategory=load_featureCategory()
wb=load_workbook('clauseTestset.xlsx')
ws=wb.active

bigFeature=[f[0] for f in featureCategory]
count=[]
for i in range(13):
    count.append([0,0])

for i in range(2,ws.max_row+1):
    feature=ws.cell(row=i,column=4).value.split(',')
    
    #仅包含单类特征的句子
    bf=[f for f in bigFeature if f in feature]
    if len(bf)==1:
        index=bigFeature.index(bf[0])
        count[index][0]+=1

    rf=list(set([featureReplace(f,featureCategory) for f in feature]))
    #if len(rf)==1:
    for f in rf:
        index=bigFeature.index(f)
        count[index][1]+=1
    ws.cell(row=i,column=7).value=','.join(rf)
wb.save('1111.xlsx')
for i in range(13):
    print(bigFeature[i],count[i][0],count[i][1])

    
    
    
