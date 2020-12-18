# coding:utf-8
from __future__ import division
import numpy as np
from openpyxl import load_workbook
from pyspark import SparkContext
from pyspark import SparkConf
from pyspark.mllib.fpm import PrefixSpan
from numpy import log
import matplotlib.pyplot as plt
import time
time1=time.time()

def load_featureSet(filename):#载入特征集
    wb=load_workbook(filename)
    ws=wb.active
    featureSet=[]
    for i in range(1,ws.max_row+1):
        feature=ws.cell(row=i,column=1).value        
        featureSet.append(feature)
    return featureSet

#导入标记好的初始评论数据
def loadDataMat(filename):
    f=open(filename,encoding='utf-8')
    dataSet=[row.strip() for row in f.readlines()]
    commentSet=[]
    labels=[]
    for i in range(1,len(dataSet)):
        data=dataSet[i].split('\t')
        comment=data[:-1]
        label=int(data[-1])
        commentSet.append(comment)
        labels.append(label)
    return commentSet,labels

def clauseCut(commentSet,labelSet):
    clauseSet=[]
    clauseLabel=[]
    for i in range(len(commentSet)):
        label=labelSet[i]
        for j in range(len(commentSet[i])):
            clause=commentSet[i][j].split(' ')
            if len(clause)>2:
                clauseSet.append(clause)
                clauseLabel.append(label)
    return clauseSet,clauseLabel

def formatDataMat(trainMat):
    dataMat=[]
    for train in trainMat:
        data=[[w] for w in train]
        dataMat.append(data)
    return dataMat

def miningPatternSet(result,m,n,featureSet):
    patternSet=[]
    for i in range(len(result)):
        if len(result[i][0])<m:continue
        if len(result[i][0])>n:continue
        pattern=[]
        for j in range(len(result[i][0])):
            pattern.append(result[i][0][j][0])
        if len(set(pattern))!=len(pattern):continue
        freq=result[i][1]
        feature=[fea for fea in featureSet if fea in pattern]
        if len(feature)==1:
            patternSet.append([pattern,feature,freq,0,0,0,0])
    return patternSet

#判断一个模式是否为另一个模式的子模式
def pattern_contain(pattern,comment):
    all_index=[]
    #首先计算出待测模式中的每个词在评论中的位置
    #print(pattern)
    for pat in pattern:
        all_index.append([i for i in range(len(comment)) if comment[i]==pat])
    position=-1
    for i in range(len(all_index)):
        if len(all_index[i])==0:return 0
        count=0
        for j in range(len(all_index[i])):
            if all_index[i][j]>position:
                position=all_index[i][j]
                break
            else:
                count+=1
        if count==len(all_index[i]):
            return 0
    return 1

def chi_square(a,b,c,d):
    if a==0 and c==0:
        return 0
    N=a+b+c+d
    return N*(a*d-b*c)**2/((a+b)*(a+c)*(c+d)*(b+d)+0.1)

def PMI_polarity(a,b,c,d):#没有使用频率的PMI算法&改进后的PMI算法
    if a==0 and c==0:return 0-0
    N=a+b+c+d
    if a==0:return 0-c*log(N*c/((c+a)*(d+c)))/log(2)
    if c==0:return a*log(N*a/((a+b)*(a+c)))/log(2)-0
    return a*log(N*a/((a+b)*(a+c)))/log(2)-c*log(N*c/((c+a)*(d+c)))/log(2)

def patternMeasure(patternSet,trainMat,trainLabel,featureSet):
    for i in range(len(trainMat)):
        #if i%1000==0:print(i)
        commentVec=trainMat[i]
        label=trainLabel[i]
        c_feature=[fea for fea in featureSet if fea in commentVec]
        for j in range(len(patternSet)):
            pattern=patternSet[j][0]
            p_feature=patternSet[j][1]
            if len(set(c_feature)&set(p_feature))>0:#相当于子数据集的筛选
                if pattern_contain(pattern,commentVec):
                    if label==1:
                        patternSet[j][3]+=1
                    if label==-1:
                        patternSet[j][5]+=1
                else:
                    if label==1:
                        patternSet[j][4]+=1
                    if label==-1:
                        patternSet[j][6]+=1
    return patternSet

def patternSelection(patternSet):
    FBIOPSet=[]
    for i in range(len(patternSet)):
        chi_square_value=chi_square(patternSet[i][3],patternSet[i][4],patternSet[i][5],patternSet[i][6])
        if chi_square_value<6.635:
            continue
        PMI_polarity_value=PMI_polarity(patternSet[i][3],patternSet[i][4],patternSet[i][5],patternSet[i][6])
        if PMI_polarity_value>0:
            polarity=1
        else:
            polarity=-1
        FBIOPSet.append([patternSet[i][0],polarity*chi_square_value])
    return FBIOPSet

def patternFilter(candidateFIO):#去除兼容模式
    FIO_patternSet=[]
    for i in range(len(candidateFIO)):
        pattern1=candidateFIO[i]
        count=0
        for j in range(len(candidateFIO)):
            if j==i:continue
            pattern2=candidateFIO[j]
            if pattern_contain(pattern1[0],pattern2[0]):
                count+=1
        if count==0:
            FIO_patternSet.append(candidateFIO[i])
    return FIO_patternSet

def senMeasure(FIO_patternSet):
    polarity=0
    for pattern in FIO_patternSet:
        polarity+=pattern[1]
    return polarity

def evaluate(testMat,testLabel,patternSet):
    labelSet=list(set(testLabel))
    results_count=np.zeros((len(labelSet),len(labelSet)))
    for i in range(len(testMat)):
        polarityTotal=0
        for j in range(len(testMat[i])):
            clauseVec=testMat[i][j].split(' ')
            candidateFIO=[]
            for k in range(len(patternSet)):
                if pattern_contain(patternSet[k][0],clauseVec):
                    candidateFIO.append(patternSet[k])
            if len(candidateFIO)==0:
                continue
            FIO_patternSet=patternFilter(candidateFIO)
            polarity=senMeasure(FIO_patternSet)
            polarityTotal+=polarity
        if polarityTotal>0:predict=1
        if polarityTotal<0:predict=-1
        if polarityTotal==0:continue
        fact=testLabel[i]
        index1=labelSet.index(fact)
        index2=labelSet.index(predict)
        results_count[index1][index2]+=1
        
    fmeasure={}
    total_TP= 0
    for idx in range(len(labelSet)):
        metric={}
        TP=results_count[idx,idx]
        total_TP += TP
        precision= TP/float(np.sum(results_count,axis=0)[idx]+0.5)
        recall= TP/float(np.sum(results_count,axis=1)[idx]+0.5)
        f_score=2*precision*recall/float(recall+precision)
        metric['p']=precision
        metric['r']=recall
        metric['f']=f_score
        fmeasure[labelSet[idx]]=metric
    accuracy=total_TP/np.sum(results_count)
    fmeasure['acc']=accuracy
    return fmeasure

def drawFigure(x,y1,y2,y3):
    plt.figure()
    plt.plot(x,y1)
    plt.plot(x,y2)
    plt.plot(x,y3)
    plt.xlabel('C')
    plt.ylabel('accuracy')
    plt.title('the accuracy of FBIOP support selection')
    plt.savefig('accuracy_FBIOP_sup.jpg')
    plt.show()

def saveToTxt(x,train,dev,test):
    f=open('accuracy_FBIOP_sup.txt','w',encoding='utf-8')
    for i in range(len(x)):
        f.write('%.4f %.4f %.4f %.4f'%(x[i],train[i],dev[i],test[i]))
        f.write('\n')
    f.close()
    
sc=SparkContext("local","testing")
print(sc)

featureSet=load_featureSet('data/featureSet.xlsx')
print('正在载入训练集')
trainMat,trainLabel=loadDataMat('mat/trainMat.txt')
print('正在载入验证集')
devMat,devLabel=loadDataMat('mat/devMat.txt')
print('正在载入测试集')
testMat,testLabel=loadDataMat('mat/testMat.txt')

print('正在格式化训练集')
clauseSet,clauseLabel=clauseCut(trainMat,trainLabel)
formatTrainMat=formatDataMat(clauseSet)

rdd=sc.parallelize(formatTrainMat,2)

best_dev_accuracy=-1
best_support=5

x=[]
train=[]
dev=[]
test=[]

for sup in range(5,50):
    print('当前支持度设置为%d'%sup)
#设置支持度
    model=PrefixSpan.train(rdd,sup/len(formatTrainMat))
    patternSet=miningPatternSet(model.freqSequences().collect(),2,100,featureSet)
    patternSet=patternMeasure(patternSet,clauseSet,clauseLabel,featureSet)
    FBIOP=patternSelection(patternSet)
    fmeasure=evaluate(trainMat,trainLabel,FBIOP)
    
    print('训练集的结果')
    accuracy1=fmeasure['acc']
    print("acc: %.4f" %(accuracy1))
    for k in fmeasure.keys():
        if k=='acc':continue
        print('label %s    p: %.4f, r: %.4f, f: %.4f'%(k, fmeasure[k]['p'], fmeasure[k]['r'], fmeasure[k]['f']))

    print('验证集的结果')
    fmeasure=evaluate(devMat,devLabel,FBIOP)
    accuracy2=fmeasure['acc']
    print("acc: %.4f" %(accuracy2))
    for k in fmeasure.keys():
        if k=='acc':continue
        print('label %s    p: %.4f, r: %.4f, f: %.4f'%(k, fmeasure[k]['p'], fmeasure[k]['r'], fmeasure[k]['f']))

    print('测试集的结果')
    fmeasure=evaluate(testMat,testLabel,FBIOP)
    accuracy3=fmeasure['acc']
    print("acc: %.4f" %(accuracy3))
    for k in fmeasure.keys():
        if k=='acc':continue
        print('label %s    p: %.4f, r: %.4f, f: %.4f'%(k, fmeasure[k]['p'], fmeasure[k]['r'], fmeasure[k]['f']))

    if accuracy2>best_dev_accuracy:
        best_dev_accuracy=accuracy2
        best_support=sup
    print(best_dev_accuracy)
    print('\n')
    x.append(sup)
    train.append(accuracy1)
    dev.append(accuracy2)
    test.append(accuracy3)

print('验证集上最佳结果:')
print(best_dev_accuracy)
model=PrefixSpan.train(rdd,best_support/len(formatTrainMat))
patternSet=miningPatternSet(model.freqSequences().collect(),2,100,featureSet)
patternSet=patternMeasure(patternSet,clauseSet,clauseLabel,featureSet)
FBIOP=patternSelection(patternSet)
fmeasure=evaluate(testMat,testLabel,FBIOP)
accuracy=fmeasure['acc']
print('测试集上最终结果：')
print("acc: %.4f" %accuracy)
for k in fmeasure.keys():
    if k=='acc':continue
    print('label %s    p: %.4f, r: %.4f, f: %.4f'%(k, fmeasure[k]['p'], fmeasure[k]['r'], fmeasure[k]['f']))
drawFigure(x,train,dev,test)
saveToTxt(x,train,dev,test)
time2=time.time()
print('一共耗时%d秒'%(time2-time1))
