# coding:utf-8
from __future__ import division
from openpyxl import load_workbook
import numpy as np
import jieba
import jieba.posseg as pseg
jieba.load_userdict('data/userdict.txt')
import re

def load_featureSet(filename):#载入特征集
    wb=load_workbook(filename)
    ws=wb.active
    featureSet=[]
    for i in range(1,ws.max_row+1):
        feature=ws.cell(row=i,column=1).value        
        featureSet.append(feature)
    return featureSet

#导入标记好的初始评论数据
def load_initComment(filename):
    f=open(filename,encoding='utf-8')
    dataSet=[row.strip() for row in f.readlines()]
    
    commentSet=[]
    labels=[]
    for i in range(1,len(dataSet)):
        data=dataSet[i].split('\t')
        comment=data[0]
        label=int(data[1])
        commentSet.append(comment)
        labels.append(label)
    return np.array(commentSet),np.array(labels)

#将长句切分为短句
def clauseSegmentation(comment):
    reg1=re.compile(u'。+')
    reg2=re.compile(u'[？?]+')
    reg3=re.compile(u'[！!]+')
    reg4=re.compile(u'[;；]+')
    reg5=re.compile(u'\d[、,，：:]+')
    reg6=re.compile(u'…+')
    reg7=re.compile(u'（\d）')
    reg9=re.compile(u'，+')
    reg10=re.compile(u'\t+')
    comment=reg1.sub(u'。【Instead】',comment)
    comment=reg2.sub(u'？【Instead】',comment)
    comment=reg3.sub(u'！【Instead】',comment)
    comment=reg4.sub(u'；【Instead】',comment)
    comment=reg5.sub(u'【Instead】',comment)
    comment=reg6.sub(u'……【Instead】',comment)
    comment=reg7.sub(u'',comment)
    comment=reg9.sub('，',comment)
    comment=reg10.sub('【Instead】',comment)
    clauseSet=re.split(u'【Instead】',comment)
    clauseList=[clause for clause in clauseSet if len(clause)>3]
    return clauseList

#分词，对词性进行筛选,并存储结果
def dataProcess(featureSet,commentSet,labelSet,filterWord,POS,filename):
    f=open(filename,'w',encoding='utf-8')
    f.write('数据集')
    for i in range(len(commentSet)):
        if i%1000==0:print(i)
        comment=commentSet[i]
        clauseList=clauseSegmentation(comment)
        data=[]
        for clause in clauseList:
            clauseVec=[w.word for w in pseg.cut(clause) if w.flag in POS if not w.word in filterWord]
            data.append(clauseVec)
        vecvec=[' '.join(d) for d in data if len(d)>1]
        lenth=sum([len(d) for d in data])
        feaCount=sum([1 for fea in featureSet if fea in comment])
        if lenth<3:continue
        if feaCount==0:continue
        f.write('\n%s\t%d'%('\t'.join(vecvec),labelSet[i]))
    f.close()

filterWord=u'，。的 ；了、：吧是'
POS=['n','d','a','v','c','i','l','ad']
featureSet=load_featureSet('data/featureSet.xlsx')
print('正在处理训练集')
reviewTrain,trainLabel=load_initComment('data/reviewTrain.txt')
dataProcess(featureSet,reviewTrain,trainLabel,filterWord,POS,'mat/trainMat.txt')
print('正在处理验证集')
reviewDev,devLabel=load_initComment('data/reviewDev.txt')
dataProcess(featureSet,reviewDev,devLabel,filterWord,POS,'mat/devMat.txt')
print('正在处理测试集')
reviewTest,testLabel=load_initComment('data/reviewTest.txt')
dataProcess(featureSet,reviewTest,testLabel,filterWord,POS,'mat/testMat.txt')
