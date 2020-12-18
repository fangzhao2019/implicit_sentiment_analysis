from __future__ import division
import jieba
import jieba.posseg as pseg
jieba.load_userdict('G:/userdict.txt')
from openpyxl import load_workbook
from openpyxl import Workbook

def calResult(a,b,c,d):
    pp=a/(a+c)
    pr=a/(a+b)
    pf=2*pp*pr/(pp+pr)
    np=d/(b+d)
    nr=d/(c+d)
    nf=2*np*nr/(np+nr)
    accuracy=(a+d)/(a+b+c+d)
    return pp,pr,pf,np,nr,nf,accuracy

wb=load_workbook('clauseResult.xlsx')
ws=wb.active
dic={}
for i in range(2,ws.max_row+1):
    feature=ws.cell(row=i,column=7).value
    label=ws.cell(row=i,column=6).value
    predictor=ws.cell(row=i,column=8).value
    if not feature in dic.keys():
        dic[feature]=[0,0,0,0]
    if label==1:
        if predictor==1:
            dic[feature][0]+=1
        if predictor==-1:
            dic[feature][1]+=1
    if label==-1:
        if predictor==1:
            dic[feature][2]+=1
        if predictor==-1:
            dic[feature][3]+=1

for key in dic.keys():
    a=dic[key][0]
    b=dic[key][1]
    c=dic[key][2]
    d=dic[key][3]
    pp,pr,pf,np,nr,nf,accuracy=calResult(a,b,c,d)
    print(key)
    print(pp,pr,pf)
    print(np,nr,nf)
    print(accuracy)
ta,tb,tc,td=0,0,0,0
for key in dic.keys():
    ta+=dic[key][0]
    tb+=dic[key][1]
    tc+=dic[key][2]
    td+=dic[key][3]
    pp,pr,pf,np,nr,nf,accuracy=calResult(ta,tb,tc,td)
print('total')
print(pp,pr,pf)
print(np,nr,nf)
print(accuracy)
for key in dic.keys():
    print(key,dic[key])
