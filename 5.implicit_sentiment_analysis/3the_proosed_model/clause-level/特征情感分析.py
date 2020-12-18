from __future__ import division
import jieba
import jieba.posseg as pseg
jieba.load_userdict('G:/userdict.txt')
from openpyxl import load_workbook
from openpyxl import Workbook

def load_featureSet(filename):#载入特征集
    wb=load_workbook(filename)
    ws=wb.active
    featureSet=[]
    for i in range(1,ws.max_row+1):
        feature=ws.cell(row=i,column=1).value        
        featureSet.append(feature)
    return featureSet

def load_patternSet(filename):
    wb=load_workbook(filename)
    ws=wb.active
    
    patternSet=[]
    for i in range(2,ws.max_row+1):
        pattern=ws.cell(row=i,column=1).value.split(',')
        feature=ws.cell(row=i,column=2).value.split(',')
        chi_square=float(ws.cell(row=i,column=4).value)
        PMI_value=ws.cell(row=i,column=7).value
        if PMI_value>0:polarity=1
        else:polarity=-1
        patternSet.append([pattern,feature,chi_square,polarity])
    return patternSet

#判断一个模式是否为另一个模式的子模式
def pattern_contain(pattern,comment):
    all_index=[]
    #首先计算出待测模式中的每个词在评论中的位置
    #print(pattern)
    for pat in pattern:
        all_index.append([i for i in range(len(comment)) if comment[i]==pat])
    position=-1
    for i in range(len(all_index)):
        if len(all_index[i])==0:return 0
        count=0
        for j in range(len(all_index[i])):
            if all_index[i][j]>position:
                position=all_index[i][j]
                break
            else:
                count+=1
        if count==len(all_index[i]):
            return 0
    return 1

def patternFilter(candidateFIO):#去除兼容模式
    FIO_patternSet=[]
    for i in range(len(candidateFIO)):
        pattern1=candidateFIO[i]
        count=0
        for j in range(len(candidateFIO)):
            if j==i:continue
            pattern2=candidateFIO[j]
            if pattern1[1]==pattern2[1] and pattern_contain(pattern1[0],pattern2[0]):
                count+=1
        if count==0:
            FIO_patternSet.append(candidateFIO[i])
    return FIO_patternSet

def senMeasure(FIO_patternSet):
    polarity=0
    for pattern in FIO_patternSet:
        polarity+=pattern[2]*pattern[3]
    if polarity>0:
        return 1
    else:
        return -1
total=0
error=0
featureSet=load_featureSet('../../featureSet.xlsx')
patternSet=load_patternSet(u'FIOP.xlsx')
wb=load_workbook('clauseTestset.xlsx')
ws=wb.active

for i in range(2,ws.max_row+1):
    if i%100==0:print(i)
    comment=ws.cell(row=i,column=1).value
    label=ws.cell(row=i,column=6).value
    commentVec=[w.word for w in pseg.cut(comment)]
    candidateFIO=[]
    for j in range(len(patternSet)):
        if pattern_contain(patternSet[j][0],commentVec):
            candidateFIO.append(patternSet[j])
    if len(candidateFIO)==0:
        ws.cell(row=i,column=8).value=0
        continue
    total+=1
    FIO_patternSet=patternFilter(candidateFIO)
    polarity=senMeasure(FIO_patternSet)
    ws.cell(row=i,column=8).value=polarity
wb.save('clauseResult.xlsx')
