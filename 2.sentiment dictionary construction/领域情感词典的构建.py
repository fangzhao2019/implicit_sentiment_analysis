# -*- coding:utf-8 -*-
from __future__ import division
import jieba.posseg as pseg
from openpyxl import load_workbook
from openpyxl import Workbook
from numpy import log
import numpy
import jieba

def load_dataSet(filename):#载入数据集
    f=open(filename,encoding='utf-8')
    dataSet=[]
    for sentence in f.readlines():
        dataSet.extend(sentence.strip())
    return dataSet

def load_commentSet(filename):#载入评论数据集并进行分词
    wb=load_workbook(filename)
    ws=wb.active
    commentSet=[]
    for i in range(2,ws.max_row+1):
        comment=ws.cell(row=i,column=1).value
        #预选的情感词点，对词性进行筛选
        data=[w.word for w in pseg.cut(comment) if w.flag=='a' or w.flag=='d' or w.flag=='v']
        commentSet.append(data)
    return commentSet

def createVocabList(dataSet):#根据评论数据集生成词汇表
    vocabSet = set([])  #create empty setv
    for document in dataSet:
        vocabSet = vocabSet | set(document) #求两个集合的并集
    return list(vocabSet)

def vocabFilter(commentSet,myVocabList,stopWord):#根据停用词和词频对词汇表进行过滤
    myVocab=[]#新的词汇表
    for vocab in myVocabList:
        if vocab in stopWord:#去除停用词
            continue
        count=0#去除低频词
        for comment in commentSet:
            if vocab in comment:
                count+=1
        if count>10:#满足要求则加入新的词汇表
            myVocab.append(vocab)
    return myVocab

def calNumberMat(filename,myVocab):
    wb=load_workbook(filename)
    ws=wb.active
    countSet=numpy.zeros((len(myVocab),4))
    for i in range(2,ws.max_row+1):
        if i%1000==0:print(i)
        comment=ws.cell(row=i,column=1).value
        label=ws.cell(row=i,column=2).value
        wordVector=[w for w in jieba.cut(comment)]
        for j in range(len(myVocab)):
            if myVocab[j] in wordVector:
                if label==1:
                    countSet[j][0]+=1
                if label==-1:
                    countSet[j][2]+=1
            else:
                if label==1:
                    countSet[j][1]+=1
                if label==-1:
                    countSet[j][3]+=1
    return countSet

def chi_square(a,b,c,d):
    if a==0 and c==0:
        return 0
    N=a+b+c+d
    return N*(a*d-b*c)**2/((a+b)*(a+c)*(c+d)*(b+d))

def PMI_polarity(a,b,c,d):#没有使用频率的PMI算法&改进后的PMI算法
    if a==0 and c==0:return 0,0
    N=a+b+c+d
    if a==0:return 0,c*log(N*c/((c+a)*(d+c)))/log(2)
    if c==0:return a*log(N*a/((a+b)*(a+c)))/log(2),0
    return a*log(N*a/((a+b)*(a+c)))/log(2),c*log(N*c/((c+a)*(d+c)))/log(2)

stopWord=load_dataSet('..\stopWord.txt')
commentSet=load_commentSet(u'0comment.xlsx')
myVocabList=createVocabList(commentSet)
print(u'载入评论成功')
myVocab=vocabFilter(commentSet,myVocabList,stopWord)
print(u'生成词语集合，共包含词语%d个'%(len(myVocab)))

countSet=calNumberMat(u'../0原始数据/合并数据.xlsx',myVocab)
print(u'正在保存结果')
new_wb=Workbook()
new_ws=new_wb.active
new_ws.cell(row=1,column=1).value=u'vocab'
new_ws.cell(row=1,column=2).value=u'flag'
new_ws.cell(row=1,column=3).value=u'chi_square'
new_ws.cell(row=1,column=4).value=u'PMI_positive'
new_ws.cell(row=1,column=5).value=u'PMI_negative'
new_ws.cell(row=1,column=6).value=u'PMI_value'
m=1
for i in range(len(countSet)):
    chi_square_value=chi_square(countSet[i][0],countSet[i][1],countSet[i][2],countSet[i][3])
    if chi_square_value<3.84:
        continue
    PMI_polarity_value1,PMI_polarity_value2=PMI_polarity(countSet[i][0],countSet[i][1],countSet[i][2],countSet[i][3])
    m+=1
    vv=[w.flag for w in pseg.cut(myVocab[i])]
    new_ws.cell(row=m,column=1).value=myVocab[i]
    new_ws.cell(row=m,column=2).value=''.join(vv[0])
    new_ws.cell(row=m,column=3).value=chi_square_value
    new_ws.cell(row=m,column=4).value=PMI_polarity_value1
    new_ws.cell(row=m,column=5).value=PMI_polarity_value2
    new_ws.cell(row=m,column=6).value=PMI_polarity_value1-PMI_polarity_value2
new_wb.save(u'情感词典.xlsx')





