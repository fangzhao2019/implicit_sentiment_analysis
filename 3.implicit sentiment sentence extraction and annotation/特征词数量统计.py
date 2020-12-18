# -*- coding: utf-8 -*-
from __future__ import division
import jieba
jieba.load_userdict('E:/博二/奇瑞项目/项目代码/userdict.txt')
from openpyxl import load_workbook
from openpyxl import Workbook

featureSet=[]
wb=load_workbook('E:\博二\奇瑞项目\项目代码\产品特征-情感词_数据集.xlsx')
ws=wb.active
for i in range(1,ws.max_row+1):
    feature=ws.cell(row=i,column=1).value
    featureSet.append(feature)

print(len(featureSet))
commentSet=[]
wb2=load_workbook('1cutResult.xlsx')
ws=wb2.active
for i in range(2,ws.max_row+1):
    comment=ws.cell(row=i,column=1).value
    commentSet.append(list(jieba.cut(comment)))

count=[0]*len(featureSet)
for i in range(len(commentSet)):
    for j in range(len(featureSet)):
        if featureSet[j] in commentSet[i]:
            count[j]+=1
for i in range(len(count)):
    print(featureSet[i],count[i])
