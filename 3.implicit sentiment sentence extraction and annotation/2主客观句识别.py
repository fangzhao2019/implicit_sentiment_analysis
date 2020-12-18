# -*- coding:utf-8 -*-
import jieba
jieba.load_userdict('G:/userdict.txt')
import jieba.posseg as pseg
import re
from openpyxl import load_workbook
from openpyxl import Workbook

def load_featureSet(filename):#载入数据集
    wb=load_workbook(filename)
    ws=wb.active
    featureSet=[]
    for i in range(1,ws.max_row+1):
        feature=ws.cell(row=i,column=1).value        
        featureSet.append(feature)
    return featureSet
    
featureSet=load_featureSet('../featureSet.xlsx')#特征集
senwordSet=load_featureSet(u'../senwordSet.xlsx')#情感词集

wb=load_workbook(u'clauseSet.xlsx')#待处理的数据
ws=wb.active
new_wb=Workbook()#创建新excel
new_ws=new_wb.active
new_ws.cell(row=1,column=1).value='comment'
new_ws.cell(row=1,column=2).value='score'
new_ws.cell(row=1,column=3).value='lenth'
new_ws.cell(row=1,column=4).value='featureNum'
new_ws.cell(row=1,column=5).value='feature'
new_ws.cell(row=1,column=6).value='senword'

m=1
for i in range(2,ws.max_row+1):
    if i%10000==0:print(i)
    comment=ws.cell(row=i,column=1).value
    score=ws.cell(row=i,column=2).value
    commentVec=[w for w in jieba.cut(comment)]
    #情感词统计
    senCount=[s for s in senwordSet if s in commentVec]
    #特征统计
    featureCount=[f for f in featureSet if f in commentVec]
    #if senCount==0:#客观句
        #if count>1:
    m+=1
    new_ws.cell(row=m,column=1).value=comment
    new_ws.cell(row=m,column=2).value=score
    new_ws.cell(row=m,column=3).value=len(commentVec)
    new_ws.cell(row=m,column=4).value=len(featureCount)
    new_ws.cell(row=m,column=5).value=','.join(featureCount)
    new_ws.cell(row=m,column=6).value=','.join(senCount)
new_wb.save(u'句子统计1.xlsx')
print(u'句子的数目为：%d个'%(m-1))

