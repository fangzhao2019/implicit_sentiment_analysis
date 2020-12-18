# -*- coding:utf-8 -*-
import jieba
jieba.load_userdict('G:/userdict.txt')
import jieba.posseg as pseg
import re
from openpyxl import load_workbook
from openpyxl import Workbook

def load_featureSet(filename):#载入数据集
    wb=load_workbook(filename)
    ws=wb.active
    featureSet=[]
    for i in range(1,ws.max_row+1):
        feature=ws.cell(row=i,column=1).value        
        featureSet.append(feature)
    return featureSet
    
featureSet=load_featureSet('../../featureSet.xlsx')#特征集

wb=load_workbook(u'clauseTestset.xlsx')#待处理的数据
ws=wb.active

for i in range(2,ws.max_row+1):
    if i%1000==0:print(i)
    comment=ws.cell(row=i,column=1).value
    commentVec=[w for w in jieba.cut(comment)]
    featureCount=[f for f in featureSet if f in commentVec]
    ws.cell(row=i,column=4).value=','.join(featureCount)
    ws.cell(row=i,column=5).value=len(featureCount)
wb.save(u'clauseTestset2.xlsx')
