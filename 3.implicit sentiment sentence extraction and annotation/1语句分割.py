# -*- coding:utf-8 -*-
from openpyxl import Workbook
from openpyxl import load_workbook
import re

wb1=load_workbook(u'0comment.xlsx')
ws1=wb1.active

wb2=Workbook()
ws2=wb2.active
ws2.cell(row=1,column=1).value='comment'
ws2.cell(row=1,column=2).value='polarity'
ws2.cell(row=1,column=3).value='lenth'

m=1
total=0
for i in range(2,ws1.max_row+1):
    comment=ws1.cell(row=i,column=1).value
    polarity=ws1.cell(row=i,column=2).value
    
    reg1=re.compile(u'。+')
    reg2=re.compile(u'[？?]+')
    reg3=re.compile(u'[！!]+')
    reg4=re.compile(u'[;；]+')
    reg5=re.compile(u'\d[、,，：:]+')
    reg6=re.compile(u'…+')
    reg7=re.compile(u'（\d）')
    reg8=re.compile(u'\s+')
    reg9=re.compile(u'，+')
    comment=reg1.sub(u'。【Instead】',comment)
    comment=reg2.sub(u'？【Instead】',comment)
    comment=reg3.sub(u'！【Instead】',comment)
    comment=reg4.sub(u'；【Instead】',comment)
    comment=reg5.sub(u'【Instead】',comment)
    comment=reg6.sub(u'……【Instead】',comment)
    comment=reg7.sub(u'',comment)
    comment=reg8.sub(u'，',comment)
    comment=reg9.sub('，',comment)
    sentence=re.split(u'【Instead】',comment)
    total+=len(sentence)
    for sen in sentence:
        lenth=len(sen.strip())
        if lenth<6 or lenth>100:
            continue
        m+=1
        ws2.cell(row=m,column=1).value=sen.strip()
        ws2.cell(row=m,column=2).value=polarity
        ws2.cell(row=m,column=3).value=lenth
print(total,m-1)
wb2.save(u'1.1cutResult.xlsx')
