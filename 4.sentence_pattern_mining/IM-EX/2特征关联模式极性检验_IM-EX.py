# -*- coding:utf-8 -*-
from __future__ import division
import jieba
jieba.load_userdict('G:/userdict.txt')
from openpyxl import load_workbook
from openpyxl import Workbook
from numpy import log
import time
time1=time.time()

def load_featureSet(filename):#载入特征集
    wb=load_workbook(filename)
    ws=wb.active
    featureSet=[]
    for i in range(1,ws.max_row+1):
        feature=ws.cell(row=i,column=1).value        
        featureSet.append(feature)
    return featureSet

def load_patternSet(filename):
    wb=load_workbook(filename)
    ws=wb.active
   
    patternSet=[]
    for i in range(2,ws.max_row+1):
        pattern=ws.cell(row=i,column=1).value
        feature=ws.cell(row=i,column=2).value
        frequency=ws.cell(row=i,column=3).value
        data=[pattern,feature,frequency,0,0,0,0]
        patternSet.append(data)
    return patternSet

#判断一个模式是否为另一个模式的子模式
def pattern_contain(pattern,comment):
    all_index=[]
    #首先计算出待测模式中的每个词在评论中的位置
    #print(pattern)
    for pat in pattern:
        all_index.append([i for i in range(len(comment)) if comment[i]==pat])
    position=-1
    for i in range(len(all_index)):
        if len(all_index[i])==0:return 0
        count=0
        for j in range(len(all_index[i])):
            if all_index[i][j]>position:
                position=all_index[i][j]
                break
            else:
                count+=1
        if count==len(all_index[i]):
            return 0
    return 1

def chi_square(a,b,c,d):
    if a==0 and c==0:
        return 0
    N=a+b+c+d
    return N*(a*d-b*c)**2/((a+b)*(a+c)*(c+d)*(b+d))

def PMI_polarity(a,b,c,d):#没有使用频率的PMI算法&改进后的PMI算法
    if a==0 and c==0:return 0,0
    N=a+b+c+d
    if a==0:return 0,c*log(N*c/((c+a)*(d+c)))/log(2)
    if c==0:return a*log(N*a/((a+b)*(a+c)))/log(2),0
    return a*log(N*a/((a+b)*(a+c)))/log(2),c*log(N*c/((c+a)*(d+c)))/log(2)

def patternMeasure():
    patternSet=load_patternSet('CandidatePattern_IM-EX_POS.xlsx')########载入模式
    print(u'载入模式成功')
    wb=load_workbook('IM-EX_dataset.xlsx')########语料库
    ws=wb.active
    for i in range(2,ws.max_row+1):
        if i%1000==0:print(i)
        comment=ws.cell(row=i,column=1).value
        label=int(ws.cell(row=i,column=2).value)
        c_feature=ws.cell(row=i,column=5).value.split(',')

        commentVec=[w for w in jieba.cut(comment)]

        for j in range(len(patternSet)):
            pattern=patternSet[j][0].split(',')
            p_feature=patternSet[j][1].split(',')
            if len(set(c_feature)&set(p_feature))>0:#相当于子数据集的筛选
                if pattern_contain(pattern,commentVec):
                    if label==1:
                        patternSet[j][3]+=1
                    if label==-1:
                        patternSet[j][5]+=1
                else:
                    if label==1:
                        patternSet[j][4]+=1
                    if label==-1:
                        patternSet[j][6]+=1
    return patternSet

def saveToExcel(patternSet):
    print(u'正在保存结果')
    new_wb=Workbook()
    new_ws=new_wb.active
    new_ws.cell(row=1,column=1).value=u'pattern'
    new_ws.cell(row=1,column=2).value=u'feature'
    new_ws.cell(row=1,column=3).value=u'frequency'
    new_ws.cell(row=1,column=4).value=u'chi_square'
    new_ws.cell(row=1,column=5).value=u'PMI_positive'
    new_ws.cell(row=1,column=6).value=u'PMI_negative'
    new_ws.cell(row=1,column=7).value=u'PMI_value'
    new_ws.cell(row=1,column=8).value=u'abs'
    new_ws.cell(row=1,column=9).value=u'len'
    new_ws.cell(row=1,column=10).value=u'featureNum'
    m=1
    for i in range(len(patternSet)):
        chi_square_value=chi_square(patternSet[i][3],patternSet[i][4],patternSet[i][5],patternSet[i][6])
        if chi_square_value<6.635:
            continue
        PMI_polarity_value1,PMI_polarity_value2=PMI_polarity(patternSet[i][3],patternSet[i][4],patternSet[i][5],patternSet[i][6])
        m+=1
        new_ws.cell(row=m,column=1).value=patternSet[i][0]
        new_ws.cell(row=m,column=2).value=patternSet[i][1]
        new_ws.cell(row=m,column=3).value=patternSet[i][2]
        new_ws.cell(row=m,column=4).value=chi_square_value
        new_ws.cell(row=m,column=5).value=PMI_polarity_value1
        new_ws.cell(row=m,column=6).value=PMI_polarity_value2
        new_ws.cell(row=m,column=7).value=PMI_polarity_value1-PMI_polarity_value2
        new_ws.cell(row=m,column=8).value=abs(PMI_polarity_value1-PMI_polarity_value2)
        new_ws.cell(row=m,column=9).value=len(patternSet[i][0].split(','))
        new_ws.cell(row=m,column=10).value=len(patternSet[i][1].split(','))
    new_wb.save('FIO_patternSet_IM-EX_POS.xlsx')
        
patternSet=patternMeasure()
saveToExcel(patternSet)
time2=time.time()
print('一共耗时%d秒'%(time2-time1))
