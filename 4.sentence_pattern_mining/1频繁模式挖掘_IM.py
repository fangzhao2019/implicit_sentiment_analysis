# -*- coding:utf-8 -*-
from __future__ import division
import jieba
import jieba.posseg as pseg
jieba.load_userdict('G:/userdict.txt')
from openpyxl import load_workbook
from openpyxl import Workbook
import numpy
from pyspark import SparkContext
from pyspark import SparkConf
from pyspark.mllib.fpm import PrefixSpan
import os
from pyltp import Segmentor

def load_featureSet(filename):#载入特征集
    wb=load_workbook(filename)
    ws=wb.active
    featureSet=[]
    for i in range(1,ws.max_row+1):
        feature=ws.cell(row=i,column=1).value        
        featureSet.append(feature)
    return featureSet

def contain_feature(featureSet,pattern):#判断模式所包含的特征
    feature=[]
    for fea in featureSet:
        if fea in pattern:
            feature.append(fea)
    return feature

def load_commentSet(filename,filterWord,method='jieba'):#载入评论数据集并进行分词
    wb=load_workbook(filename)
    ws=wb.active
    commentSet=[]
    POS=['n','d','a','v','c','i','l','ad']######使用词性过滤
    
    if method=='jieba':
        for i in range(2,ws.max_row+1):
            comment=ws.cell(row=i,column=1).value
            #data=[[w.word] for w in pseg.cut(comment) if not w.word in filterWord]#############
            data=[[w.word] for w in pseg.cut(comment) if w.flag in POS]
            commentSet.append(data)
        print(u'数据加载成功')

    if method=='pyltp':
        LTP_DATA_DIR = 'G:\ltp_data_v3.4.0'
        cws_model_path = os.path.join(LTP_DATA_DIR, 'cws.model') 
        segmentor = Segmentor() 
        segmentor.load_with_lexicon(cws_model_path, 'G:/userdict.txt')
        for i in range(2,ws.max_row+1):
            comment=ws.cell(row=i,column=1).value
            data=[[w] for w in segmentor.segment(comment) if w not in filterWord]
            commentSet.append(data)
        print(u'数据加载成功')
        segmentor.release()  
    return commentSet

def convert(result,m,n,featureSet):
    wb=Workbook()
    ws=wb.active
    ws.cell(row=1,column=1).value='pattern'
    ws.cell(row=1,column=2).value='feature'
    ws.cell(row=1,column=3).value='freq'
    ws.cell(row=1,column=4).value='len'
    k=1
    for i in range(len(result)):
        if len(result[i][0])<m:continue
        if len(result[i][0])>n:continue
        pattern=[]
        for j in range(len(result[i][0])):
            pattern.append(result[i][0][j][0])
        if len(set(pattern))<2:continue
        freq=result[i][1]
        feature=contain_feature(featureSet,pattern)
        if len(feature)>-1:#####使用特征进行过滤
            k+=1
            if k%10000==0:print(k)
            ws.cell(row=k,column=1).value=','.join(pattern)
            #ws.cell(row=k,column=2).value=','.join(list(set(pattern)))
            #这里是否需要去除重复词来提高挖掘的效率值得探究，如果使用set，会自动排序导致错误
            ws.cell(row=k,column=2).value=','.join(feature)
            ws.cell(row=k,column=3).value=freq
            ws.cell(row=k,column=4).value=len(set(pattern))
    wb.save('IM_FrequentPattern.xlsx')

filterWord=u'，。的 ；了、：吧'
featureSet=load_featureSet('../featureSet.xlsx')
commentSet=load_commentSet(u'clauseSet.xlsx',filterWord,method='jieba')
sc=SparkContext("local","testing")
print(sc)
rdd=sc.parallelize(commentSet,2)
#设置支持度
model=PrefixSpan.train(rdd,15/len(commentSet))
#设置频繁项规模
convert(model.freqSequences().collect(),2,100,featureSet)
